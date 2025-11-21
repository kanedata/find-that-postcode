"""
Import commands for the register of geographic codes and code history database
"""

import codecs
import csv
import datetime
import io
import zipfile
from collections import defaultdict

import click
import requests
import requests_cache
import tqdm
from openpyxl import load_workbook

from findthatpostcode.commands.utils import bulk_upload, get_latest_geoportal_url
from findthatpostcode.db import get_es
from findthatpostcode.metadata import ENTITIES
from findthatpostcode.settings import DEBUG, ENTITY_INDEX

PRD_RGC = "PRD_RGC"
PRD_CHD = "PRD_CHD"
MSOA_2011_URL = (
    "https://houseofcommonslibrary.github.io/msoanames/MSOA-Names-Latest.csv"
)
MSOA_2021_URL = (
    "https://houseofcommonslibrary.github.io/msoanames/MSOA-Names-Latest2.csv"
)
DEFAULT_ENCODING = "utf-8-sig"


def process_date(value, date_format="%d/%m/%Y"):
    if value in ["", "n/a"]:
        return None
    if isinstance(value, datetime.datetime):
        return value
    return datetime.datetime.strptime(value, date_format)


def process_str(value):
    if value in ["", "n/a"]:
        return None
    if not isinstance(value, str):
        return value
    return value.strip()


def process_int(value):
    if value in ["", "n/a"]:
        return None
    if not isinstance(value, str):
        return value
    value = value.replace(",", "")
    return int(value)


def process_float(value):
    if value in ["", "n/a"]:
        return None
    if not isinstance(value, str):
        return value
    value = value.replace(",", "")
    return float(value)


@click.command("rgc")
@click.option("--url", default=None)
@click.option("--es-index", default=ENTITY_INDEX)
def import_rgc(url=None, es_index=ENTITY_INDEX):
    if DEBUG:
        requests_cache.install_cache()

    es = get_es()
    if not url:
        url = get_latest_geoportal_url(PRD_RGC)

    r = requests.get(url, stream=True)
    z = zipfile.ZipFile(io.BytesIO(r.content))
    for f in z.namelist():
        if not f.endswith(".xlsx"):
            click.echo(f"[entities] Skipping {f}")
            continue
        with z.open(f, "r") as infile:
            wb = load_workbook(
                io.BytesIO(infile.read()), read_only=True, data_only=True
            )
            sheet = wb.active
            entities: list[dict] = []
            headers = []
            if sheet is None:
                raise ValueError("No active sheet found in RGC workbook")
            for row in sheet.iter_rows(values_only=True):
                if not headers:
                    headers = [h.strip() if isinstance(h, str) else str(h) for h in row]
                    continue

                entity = dict(zip(headers, row))
                if not entity["Entity code"]:
                    continue

                # tidy up a couple of records
                if isinstance(entity["Related entity codes"], str):
                    related_codes = (
                        entity["Related entity codes"].replace("n/a", "").split(", ")
                    )
                else:
                    related_codes = []

                entities.append(
                    {
                        "_index": es_index,
                        "_type": "_doc",
                        "_op_type": "update",
                        "_id": entity["Entity code"],
                        "doc_as_upsert": True,
                        "doc": {
                            "code": entity["Entity code"],
                            "name": entity["Entity name"],
                            "abbreviation": entity["Entity abbreviation"],
                            "theme": entity["Entity theme"],
                            "coverage": entity["Entity coverage"],
                            "related_codes": related_codes,
                            "status": entity["Status"],
                            "live_instances": process_int(
                                entity["Number of live instances"]
                            ),
                            "archived_instances": process_int(
                                entity["Number of archived instances"]
                            ),
                            "crossborder_instances": process_int(
                                entity["Number of cross-border instances"]
                            ),
                            "last_modified": process_date(
                                entity["Date of last instance change"]
                            ),
                            "current_code_first": entity[
                                "Current code (first in range)"
                            ],
                            "current_code_last": entity["Current code (last in range)"],
                            "reserved_code": entity["Reserved code (for CHD use)"],
                            "owner": entity.get("Entity owner"),
                            "date_introduced": process_date(
                                entity["Date entity introduced on RGC"]
                            ),
                            "date_start": process_date(entity["Entity start date"]),
                            "type": ENTITIES.get(entity["Entity code"]),
                        },
                    }
                )

            bulk_upload(entities, es, es_index, "entities")


@click.command("chd")
@click.option("--url", default=None)
@click.option("--es-index", default=AREA_INDEX)
@click.option("--encoding", default=DEFAULT_ENCODING)
def import_chd(url=None, es_index=AREA_INDEX, encoding=DEFAULT_ENCODING):
    if DEBUG:
        requests_cache.install_cache()

    if not url:
        url = get_latest_geoportal_url(PRD_CHD)

    es = get_es()

    r = requests.get(url, stream=True)
    z = zipfile.ZipFile(io.BytesIO(r.content))

    areas_cache = defaultdict(list)
    areas = {}

    change_history = None
    changes = None
    equivalents = None
    for f in z.namelist():
        if f.lower().startswith("changehistory") and f.lower().endswith(".csv"):
            change_history = f
        elif f.lower().startswith("changes") and f.lower().endswith(".csv"):
            changes = f
        elif f.lower().startswith("equivalents") and f.lower().endswith(".csv"):
            equivalents = f

    if change_history is None or changes is None or equivalents is None:
        raise ValueError("Required CSV files not found in the ZIP archive")

    encodings_to_try = ["utf-8-sig", "cp1252", "windows-1252"]
    if encoding not in encodings_to_try:
        encodings_to_try.insert(0, encoding)

    for enc in encodings_to_try:
        try:
            with z.open(change_history, "r") as infile:
                click.echo("Opening {} with {}".format(infile.name, enc))
                reader = csv.DictReader(io.TextIOWrapper(infile, encoding=enc))
                for k, area in tqdm.tqdm(enumerate(reader)):
                    areas_cache[area["GEOGCD"]].append(
                        {
                            "code": area["GEOGCD"],
                            "name": area["GEOGNM"],
                            "name_welsh": area["GEOGNMW"] if area["GEOGNMW"] else None,
                            "statutory_instrument_id": area["SI_ID"]
                            if area["SI_ID"]
                            else None,
                            "statutory_instrument_title": area["SI_TITLE"]
                            if area["SI_TITLE"]
                            else None,
                            "date_start": process_date(
                                area["OPER_DATE"][0:10].split(" ")[0], "%d/%m/%Y"
                            ),
                            "date_end": process_date(
                                area["TERM_DATE"][0:10].split(" ")[0], "%d/%m/%Y"
                            ),
                            "parent": area["PARENTCD"] if area["PARENTCD"] else None,
                            "entity": process_str(area["ENTITYCD"]),
                            "owner": area["OWNER"],
                            "active": area["STATUS"] == "live",
                            "areaehect": process_float(area["AREAEHECT"]),
                            "areachect": process_float(area["AREACHECT"]),
                            "areaihect": process_float(area["AREAIHECT"]),
                            "arealhect": process_float(area["AREALHECT"]),
                            "sort_order": area["GEOGCD"],
                            "predecessor": [],
                            "successor": [],
                            "equivalents": {},
                            "type": ENTITIES.get(process_str(area["ENTITYCD"])),
                        }
                    )
            break
        except UnicodeDecodeError:
            click.echo(
                f"Failed to read {change_history} with encoding {enc}, trying next..."
            )

    for area_code, area_history in tqdm.tqdm(areas_cache.items()):
        if len(area_history) == 1:
            area = area_history[0]
            area["alternative_names"] = []
            if area["name"]:
                area["alternative_names"].append(area["name"])
            if area["name_welsh"]:
                area["alternative_names"].append(area["name_welsh"])
        else:
            area_history = sorted(area_history, key=lambda x: x["date_start"])
            area = area_history[-1]
            area["date_start"] = area_history[0]["date_start"]
            area["alternative_names"] = []
            for h in area_history:
                if h["name"] and h["name"] not in area["alternative_names"]:
                    area["alternative_names"].append(h["name"])
                if h["name_welsh"] and h["name_welsh"] not in area["alternative_names"]:
                    area["alternative_names"].append(h["name_welsh"])
        areas[area_code] = {
            "_index": es_index,
            "_type": "_doc",
            "_op_type": "update",
            "_id": area_code,
            "doc_as_upsert": True,
            "doc": area,
        }

    for enc in encodings_to_try:
        try:
            with z.open(changes, "r") as infile:
                click.echo("Opening {} with {}".format(infile.name, enc))
                reader = csv.DictReader(io.TextIOWrapper(infile, encoding=enc))
                for k, area in tqdm.tqdm(enumerate(reader)):
                    if area["GEOGCD_P"] == "":
                        continue
                    if area["GEOGCD"] in areas:
                        areas[area["GEOGCD"]]["doc"]["predecessor"].append(
                            area["GEOGCD_P"]
                        )
                    if area["GEOGCD_P"] in areas:
                        areas[area["GEOGCD_P"]]["doc"]["successor"].append(
                            area["GEOGCD"]
                        )
            break
        except UnicodeDecodeError:
            click.echo(f"Failed to read {changes} with encoding {enc}, trying next...")

    equiv = {
        "ons": ["GEOGCDO", "GEOGNMO"],
        "mhclg": ["GEOGCDD", "GEOGNMD"],
        "nhs": ["GEOGCDH", "GEOGNMH"],
        "scottish_government": ["GEOGCDS", "GEOGNMS"],
        "welsh_government": ["GEOGCDWG", "GEOGNMWG", "GEOGNMWWG"],
    }
    for enc in encodings_to_try:
        try:
            with z.open(equivalents, "r") as infile:
                click.echo("Opening {} with {}".format(infile.name, enc))
                reader = csv.DictReader(io.TextIOWrapper(infile, encoding=enc))
                for area in tqdm.tqdm(reader):
                    if "GEOGCD" not in area:
                        raise ValueError("No GEOGCD field in equivalents file")
                    if area["GEOGCD"] not in areas:
                        continue
                    for k, v in equiv.items():
                        if area[v[0]]:
                            areas[area["GEOGCD"]]["doc"]["equivalents"][k] = area[v[0]]
            break
        except UnicodeDecodeError:
            click.echo(
                f"Failed to read {equivalents} with encoding {enc}, trying next..."
            )

    bulk_upload(list(areas.values()), es, es_index, "areas")


@click.command("msoanames")
@click.option("--url", default=MSOA_2021_URL)
@click.option("--es-index", default=AREA_INDEX)
def import_msoa_names(url=MSOA_2021_URL, es_index=AREA_INDEX):
    if DEBUG:
        requests_cache.install_cache()

    es = get_es()

    r = requests.get(url, stream=True)

    reader = csv.DictReader(codecs.iterdecode(r.iter_lines(), "utf-8-sig"))
    area_updates = []
    for k, area in tqdm.tqdm(enumerate(reader)):
        areacode = None
        name = None
        name_welsh = None
        if "msoa21cd" in area.keys():
            areacode = area.get("msoa21cd")
            name = area.get("msoa21hclnm")
            name_welsh = area.get("msoa21hclnmw")
        if "msoa11cd" in area.keys():
            areacode = area.get("msoa11cd")
            name = area.get("msoa11hclnm")
            name_welsh = area.get("msoa11hclnmw")

        alt_names = [name]
        if name_welsh:
            alt_names.append(name_welsh)
        new_doc = {
            "doc": {
                "name": name,
                "name_welsh": name_welsh if name_welsh else None,
                "alternative_names": alt_names,
            }
        }
        area_updates.append(
            {
                "_index": es_index,
                "_type": "_doc",
                "_op_type": "update",
                "_id": areacode,
                "doc_as_upsert": True,
                **new_doc,
            }
        )

    bulk_upload(area_updates, es, es_index, "areas")
