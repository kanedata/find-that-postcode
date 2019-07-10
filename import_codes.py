import csv
import zipfile
import io
import argparse
import os
from datetime import datetime

from elasticsearch import Elasticsearch
from elasticsearch.helpers import bulk

from openpyxl import load_workbook

from metadata import NAME_FILES

ENTITY_INDEX = "geo_entity"
AREA_INDEX = "geo_area"

def main():
    parser = argparse.ArgumentParser(description='Import geographic codes into elasticsearch.')

    # Files to load data from
    parser.add_argument('--rgc', type=str,
                        default='data/Register_of_Geographic_codes.zip',
                        help='ZIP file for Register of Geographic Codes')
                    
    # from https://www2.gov.scot/Topics/Statistics/sns/SNSRef/StanGeoCodeRegister
    parser.add_argument('--scot', type=str,
                        default='data/Standard_Geography_Code_Register.xlsx',
                        help='XLSX file for Standard Geography Code Register (Scotland)')

    # elasticsearch options
    parser.add_argument('--es-host', default="localhost", help='host for the elasticsearch instance')
    parser.add_argument('--es-port', default=9200, help='port for the elasticsearch instance')
    parser.add_argument('--es-url-prefix', default='', help='Elasticsearch url prefix')
    parser.add_argument('--es-use-ssl', action='store_true', help='Use ssl to connect to elasticsearch')

    args = parser.parse_args()

    postcodes = []

    es = Elasticsearch(host=args.es_host, port=args.es_port, url_prefix=args.es_url_prefix, use_ssl=args.es_use_ssl)

    potential_env_vars = [
        "ELASTICSEARCH_URL",
        "ES_URL",
        "BONSAI_URL"
    ]
    for e_v in potential_env_vars:
        if os.environ.get(e_v):
            es = Elasticsearch(os.environ.get(e_v))
            break

    instance_count = 0
    with zipfile.ZipFile(args.rgc) as rgczip:
        for f in rgczip.filelist:
            if f.filename.endswith(".xlsx"):
                print("[codes] Opening %s" % f.filename)
                with rgczip.open(f, 'r') as rgcfile:
                    wb = load_workbook(rgcfile)

                    # get entity codes
                    headers = None
                    entities = {}
                    for row in wb["RGC"].values:
                        if not headers:
                            headers = row
                            continue
                        
                        entity = dict(zip(headers, row))

                        # tidy up a couple of records
                        entity["Related entity codes"] = entity["Related entity codes"].replace("n/a", "").split(", ")
                        for f in ["Date of last instance change", "Date entity introduced on RGC", "Entity start date"]:
                            if entity[f] == "n/a":
                                entity[f] = None
                            elif not isinstance(entity[f], datetime):
                                entity[f] = None

                        entities[entity["Entity code"]] = {
                            "_index": ENTITY_INDEX,
                            "_type": "_doc",
                            "_op_type": "index",
                            "_id": entity["Entity code"],
                            "code": entity['Entity code'],
                            "name": entity['Entity name'],
                            "abbreviation": entity['Entity abbreviation'],
                            "theme": entity['Entity theme'],
                            "coverage": entity['Entity coverage'],
                            "related_codes": entity['Related entity codes'],
                            "status": entity['Status'],
                            "live_instances": entity['Number of live instances'],
                            "archived_instances": entity['Number of archived instances'],
                            "crossborder_instances": entity['Number of cross-border instances'],
                            "last_modified": entity['Date of last instance change'],
                            "current_code_first": entity['Current code (first in range)'],
                            "current_code_last": entity['Current code (last in range)'],
                            "reserved_code": entity['Reserved code (for CHD use)'],
                            "owner": entity['Entity owner'],
                            "date_introduced": entity['Date entity introduced on RGC'],
                            "date_start": entity['Entity start date'],
                        }

                    print("[entities] Processed %s entities" % len(entities))
                    print("[elasticsearch] %s entities to save" % len(entities))
                    results = bulk(es, list(entities.values()))
                    print("[elasticsearch] saved %s entities to %s index" % (results[0], ENTITY_INDEX))
                    print("[elasticsearch] %s errors reported" % len(results[1]))

                    # get entity instances
                    for ws in wb:
                        if ws.title in ['RGC', 'Metadata_for_geography_listings', 'For_Scotland']:
                            continue
                            
                        headers = None
                        instances = []
                        for row in ws.values:
                            if not headers:
                                headers = row
                                continue
                            
                            instance = dict(zip(headers, row))
                            instance_count += 1
                        
                            instances.append({
                                "_index": AREA_INDEX,
                                "_type": "_doc",
                                "_op_type": "index",
                                "_id": instance["GEOGCD"],
                                "code": instance["GEOGCD"], # Nine digit code for the Instance (i.e. E01000001)
                                "name": instance["GEOGNM"], # Name of the Instance
                                "name_welsh": instance["GEOGNMW"], # Welsh name of the Instance
                                "statutory_instrument_id": instance["SI_ID"], # Number of legislation which defines the Instance, if applicable (i.e.  Statutory Instruments)
                                "statutory_instrument_title": instance["SI_TITLE"], # Name of legislation which defines the Instance, if applicable (i.e. Statutory Instruments)
                                "date_start": instance["OPER_DATE"], # Date when the Instance came into use (usually the enforcement date of legislation which defines the instance)
                                "date_end": instance["TERM_DATE"], # Date when the Instance was archived (usually the day before the enforcement date of legislation which defines the new instance)
                                "parent": instance["PARENTCD"], # Nine digit code of the parent instance in the hierarchy (if there is one)
                                "entity": instance["ENTITYCD"], # Three digit code prefix for the Entity (i.e. E01)
                                "owner": instance["OWNER"], # Geography owner of the entity
                                "active": instance["STATUS"] == "live", # States whether the Instance is live or terminated
                                "areaehect": instance.get("AREAEHECT"), # ‘Extent of the Realm’ measurement, in hectares, to 2 decimal places. ‘Extent of the Realm’ is typically co-incident with Mean Low Water
                                "areachect": instance.get("AREACHECT"), # Area to Mean High Water, in hectares, to 2 decimal places. Measurements are limited to the Mean High Water mark and include all tracts of inland water
                                "areaihect": instance.get("AREAIHECT"), # Area of inland water, in hectares, to 2 decimal places. This is the surface area of inland water with a surface area measurement of more than 1km2
                                "arealhect": instance.get("AREALHECT"), # Area to Mean High Water excluding area of inland water (land area), in hectares, to 2 decimal places.
                                "sort_order": instance["GEOGCD"],
                                "type": entities.get(instance["ENTITYCD"], {}).get("abbreviation", instance["ENTITYCD"]),
                            })

                            if len(instances)>=10000:
                                print("[entities] Processed %s instances" % instance_count)
                                print("[elasticsearch] %s instances to save" % len(instances))
                                results = bulk(es, instances)
                                print("[elasticsearch] saved %s instances to %s index" % (results[0], AREA_INDEX))
                                print("[elasticsearch] %s errors reported" % len(results[1]))
                                instances = []

                        print("[entities] Processed %s instances" % instance_count)
                        print("[elasticsearch] %s instances to save" % len(instances))
                        results = bulk(es, instances)
                        print("[elasticsearch] saved %s instances to %s index" % (results[0], AREA_INDEX))
                        print("[elasticsearch] %s errors reported" % len(results[1]))
                        
    if args.scot:
        print("[codes] Opening %s" % args.scot)
        wb = load_workbook(args.scot)

        # get entity codes
        headers = None
        entities = {}
        for row in wb["Entities"].values:
            if not headers:
                headers = row
                continue
            
            entity = dict(zip(headers, row))

            # tidy up a couple of records
            if entity["RelatedEntity"]:
                entity["RelatedEntity"] = entity["RelatedEntity"].replace("n/a", "").split(", ")
            else:
                entity["RelatedEntity"] = []
            for f in ["DateEnacted", "DateArchived", "DateLastChange"]:
                if entity[f] == "n/a":
                    entity[f] = None
                elif not isinstance(entity[f], datetime):
                    entity[f] = None
            for f in ["NumLiveInstances", "NumArchInstances"]:
                if entity[f] == "n/a":
                    entity[f] = None
                elif not isinstance(entity[f], (int,float)):
                    entity[f] = None

            entities[entity["EntityCode"]] = {
                "_index": ENTITY_INDEX,
                "_type": "_doc",
                "_op_type": "index",
                "_id": entity["EntityCode"],
                "code": entity['EntityCode'],
                "name": entity['EntityName'],
                "abbreviation": entity['EntityAcronym'],
                "theme": entity['EntityTheme'],
                "coverage": entity['EntityCoverage'],
                "related_codes": entity['RelatedEntity'],
                "status": entity['Status'],
                "live_instances": entity['NumLiveInstances'],
                "archived_instances": entity['NumArchInstances'],
                "crossborder_instances": None,
                "last_modified": entity['DateLastChange'],
                "current_code_first": entity['CodeRangeFirst'],
                "current_code_last": entity['CodeRangeLast'],
                "reserved_code": entity['ReservedCode'],
                "owner": entity['EntityOwner'],
                "date_introduced": entity['DateEnacted'],
                "date_start": entity['DateEnacted'],
                "date_archived": entity['DateArchived'],
            }

        print("[entities] Processed %s entities" % len(entities))
        print("[elasticsearch] %s entities to save" % len(entities))
        results = bulk(es, list(entities.values()))
        print("[elasticsearch] saved %s entities to %s index" % (results[0], ENTITY_INDEX))
        print("[elasticsearch] %s errors reported" % len(results[1]))

        # get entity instances
        for ws in wb:
            if ws.title in ['Entities', 'Metadata']:
                continue
                
            headers = None
            instances = []
            for row in ws.values:
                if not headers:
                    headers = row
                    continue
                
                instance = dict(zip(headers, row))
                instance_count += 1

                if instance["InstanceCode"] is None or instance["InstanceCode"] == "Pending" or instance["InstanceCode"][0] != "S":
                    continue
                    
                for f in ["DateEnacted", "DateArchived"]:
                    if instance[f] == "n/a":
                        instance[f] = None
                    elif not isinstance(instance[f], datetime):
                        instance[f] = None
            
                instances.append({
                    "_index": AREA_INDEX,
                    "_type": "_doc",
                    "_op_type": "index",
                    "_id": instance["InstanceCode"],
                    "code": instance["InstanceCode"], # Nine digit code for the Instance (i.e. E01000001)
                    "name": instance["InstanceName"], # Name of the Instance
                    "name_welsh": None, # Welsh name of the Instance
                    "statutory_instrument_id": None, # Number of legislation which defines the Instance, if applicable (i.e.  Statutory Instruments)
                    "statutory_instrument_title": instance["LegalDocOpen"], # Name of legislation which defines the Instance, if applicable (i.e. Statutory Instruments)
                    "notes": instance["NotesOpen"], # Name of legislation which defines the Instance, if applicable (i.e. Statutory Instruments)
                    "date_start": instance["DateEnacted"], # Date when the Instance came into use (usually the enforcement date of legislation which defines the instance)
                    "date_end": instance["DateArchived"], # Date when the Instance was archived (usually the day before the enforcement date of legislation which defines the new instance)
                    "parent": None, # Nine digit code of the parent instance in the hierarchy (if there is one)
                    "entity": instance["EntityCode"], # Three digit code prefix for the Entity (i.e. E01)
                    "owner": None, # Geography owner of the entity
                    "active": ((instance["Status"] is not None) and (instance["Status"].lower() == "live")), # States whether the Instance is live or terminated
                    "areaehect": None, # ‘Extent of the Realm’ measurement, in hectares, to 2 decimal places. ‘Extent of the Realm’ is typically co-incident with Mean Low Water
                    "areachect": None, # Area to Mean High Water, in hectares, to 2 decimal places. Measurements are limited to the Mean High Water mark and include all tracts of inland water
                    "areaihect": None, # Area of inland water, in hectares, to 2 decimal places. This is the surface area of inland water with a surface area measurement of more than 1km2
                    "arealhect": None, # Area to Mean High Water excluding area of inland water (land area), in hectares, to 2 decimal places.
                    "sort_order": instance["InstanceCode"],
                    "type": entities.get(instance["EntityCode"], {}).get("abbreviation", instance["EntityCode"]),
                })

                if len(instances)>=10000:
                    print("[entities] Processed %s instances" % instance_count)
                    print("[elasticsearch] %s instances to save" % len(instances))
                    results = bulk(es, instances)
                    print("[elasticsearch] saved %s instances to %s index" % (results[0], AREA_INDEX))
                    print("[elasticsearch] %s errors reported" % len(results[1]))
                    instances = []

            print("[entities] Processed %s instances" % instance_count)
            print("[elasticsearch] %s instances to save" % len(instances))
            results = bulk(es, instances)
            print("[elasticsearch] saved %s instances to %s index" % (results[0], AREA_INDEX))
            print("[elasticsearch] %s errors reported" % len(results[1]))
            

if __name__ == '__main__':
    main()
